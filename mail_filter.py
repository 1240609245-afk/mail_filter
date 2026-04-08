import json
import re
import csv
import poplib
import email
import os
import html
import time
import random
import shutil
from email.header import decode_header, make_header
from datetime import datetime, timedelta
from email.utils import parsedate_to_datetime
from dotenv import load_dotenv

# 解决部分邮件 "line too long" 问题
poplib._MAXLINE = 1024 * 1024

# 强制覆盖旧环境变量
load_dotenv(override=True)


# =========================
# 1) 风险关键词库
# =========================
IMPORTANT_LIKE_RULES = [
    "%Your Amazon-Fulfilled Inventory - Action Required%",
    "付款尝试失败",
    "%Your Amazon selling privileges have been removed%",
    "%Your seller account%",
    "%Formal Notice - Immediate Action Required%",
    "%Important%Amazon%seller%account%",
    "%Your funds transfer has been cancelled%",
    "%Seller Account Protection%",
    "%Review of your Amazon account%",
    "%KYC%",
    "%Important message about your Amazon seller account%",
    "%Action required: Important information about your product listings%",
    "%Your Amazon seller detail pages have been temporarily removed%",
    "%Listings deactivated due to potential pricing error%",
    "%Policy Warning%",
    "%Notification of Product Removal%",
    "%Notification of Restricted Products Removal%",
    "%Your Amazon.com selling account%",
    "%Notice warning%",
    "%Plaintiff%",
    "%TRO%",
    "%court%",
    "%Important information about your Stranded Fulfillment by Amazon inventory%",
    "%Automated unfulfillable removal notification%",
    "%Stranded SKUs scheduled for automatic removal%",
    "%Your amazon.com Inventory%",
    "%submit product safety documentation%",
    "%avoid listing deactivation%",
    "%reactivate listings%",
    "%listing deactivation%",
    "%listing deactivated%",
    "%listing removed%",
    "%removed%",
    "%product safety%",
    "%safety documentation%",
    "%compliance%",
    "%compliance request%",
    "%verification required%",
    "%document required%",
    "%action required%",
    "%performance notification%",
    "%intellectual property%",
    "%infringement%",
    "%rights owner%",
    "%unsafe%",
    "%hazard%",
    "%recall%",
    "%appeal%",
    "%deactivated%",
    "%suppressed%",
    "%angebote wurden aus%",
    "%entfernt%",
    "%genehmigt wird%",
]


# =========================
# 2) 分类规则
# =========================
CATEGORY_RULES = {
    "店铺绩效类": [
        "your amazon selling privileges have been removed",
        "your seller account",
        "your amazon.com selling account",
        "seller account protection",
        "review of your amazon account",
        "account review",
        "important message about your amazon seller account",
        "formal notice - immediate action required",
    ],
    "KYC审核类": [
        "kyc",
        "verification required",
        "identity verification",
        "document required",
    ],
    "付款/资金类": [
        "付款尝试失败",
        "your funds transfer has been cancelled",
        "your payment is on the way",
        "ihre auszahlung wird ausgeführt",
        "tu pago está en camino",
        "twoja płatność jest już w drodze",
        "auszahlung",
        "payment",
        "pago",
        "płatność",
        "funds transfer",
        "disbursement",
    ],
    "产品链接/合规类": [
        "your amazon-fulfilled inventory - action required",
        "action required: important information about your product listings",
        "your amazon seller detail pages have been temporarily removed",
        "listings deactivated due to potential pricing error",
        "notification of product removal",
        "notification of restricted products removal",
        "submit product safety documentation",
        "avoid listing deactivation",
        "reactivate listings",
        "listing deactivation",
        "listing deactivated",
        "listing removed",
        "angebote wurden aus",
        "entfernt",
        "genehmigt wird",
        "product safety",
        "safety documentation",
        "compliance",
        "policy warning",
        "notice warning",
        "notice: policy warning",
    ],
    "法务/侵权类": [
        "plaintiff",
        "tro",
        "court",
        "rights owner",
        "infringement",
        "intellectual property",
    ],
    "货件/库存类": [
        "important information about your stranded fulfillment by amazon inventory",
        "automated unfulfillable removal notification",
        "stranded skus scheduled for automatic removal",
        "your amazon.com inventory",
        "fba removal order",
        "removal order",
    ],
}


# =========================
# 3) 基础工具函数
# =========================
def like_to_regex(pattern: str) -> re.Pattern:
    p = pattern.strip()
    if "%" not in p:
        p = f"%{p}%"
    p_escaped = re.escape(p).replace(r"\%", ".*")
    return re.compile(p_escaped, re.IGNORECASE | re.DOTALL)


RULE_REGEX = [like_to_regex(p) for p in IMPORTANT_LIKE_RULES]


def load_accounts_from_env():
    accounts = []

    for i in range(1, 20):
        email_addr = os.getenv(f"EMAIL_{i}", "")
        password = os.getenv(f"PASSWORD_{i}", "")
        host = os.getenv(f"HOST_{i}", "")

        email_addr = email_addr.strip()
        password = password.strip()
        host = host.strip()

        if not email_addr and not password and not host:
            continue

        if not (email_addr and password and host):
            print(f"警告：第 {i} 组邮箱配置不完整，已跳过")
            continue

        accounts.append({
            "name": f"account_{i}",
            "host": host,
            "user": email_addr,
            "password": password,
        })

    if not accounts:
        raise SystemExit("没有读取到任何邮箱配置，请检查本地 .env 是否已设置 EMAIL_x / PASSWORD_x / HOST_x")

    return accounts


def decode_mime(s):
    if not s:
        return ""
    try:
        return str(make_header(decode_header(s)))
    except Exception:
        return str(s)


def get_text_headers(msg):
    subject = decode_mime(msg.get("Subject", ""))
    from_ = decode_mime(msg.get("From", ""))
    date_ = decode_mime(msg.get("Date", ""))
    return subject, from_, date_


def html_to_text_keep_tables(html_text: str) -> str:
    if not html_text:
        return ""

    text = html_text

    text = re.sub(r"(?is)<script[\s\S]*?</script>", "", text)
    text = re.sub(r"(?is)<style[\s\S]*?</style>", "", text)

    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"(?i)</p\s*>", "\n", text)
    text = re.sub(r"(?i)</div\s*>", "\n", text)
    text = re.sub(r"(?i)</li\s*>", "\n", text)
    text = re.sub(r"(?i)<li\s*>", "- ", text)

    # 表格结构保留一些间距
    text = re.sub(r"(?i)</tr\s*>", "\n", text)
    text = re.sub(r"(?i)</td\s*>", "    ", text)
    text = re.sub(r"(?i)</th\s*>", "    ", text)

    text = re.sub(r"<[^>]+>", " ", text)

    text = html.unescape(text)
    text = re.sub(r"\r", "", text)
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]{2,}", " ", text)

    return text.strip()


def safe_decode_payload(payload, charset):
    if not payload:
        return ""

    charset_candidates = []
    if charset:
        charset_candidates.append(charset)

    charset_candidates.extend([
        "utf-8",
        "gb18030",
        "gbk",
        "big5",
        "latin1",
    ])

    for cs in charset_candidates:
        try:
            return payload.decode(cs, errors="replace")
        except Exception:
            continue

    return payload.decode("utf-8", errors="replace")


def sanitize_html_for_display(raw_html: str) -> str:
    if not raw_html:
        return ""

    cleaned = raw_html

    # 去掉危险内容
    cleaned = re.sub(r"(?is)<script[\s\S]*?</script>", "", cleaned)
    cleaned = re.sub(r"(?is)<style[\s\S]*?</style>", "", cleaned)
    cleaned = re.sub(r"(?is)<iframe[\s\S]*?</iframe>", "", cleaned)
    cleaned = re.sub(r"(?is)<object[\s\S]*?</object>", "", cleaned)
    cleaned = re.sub(r"(?is)<embed[\s\S]*?</embed>", "", cleaned)
    cleaned = re.sub(r"(?is)<form[\s\S]*?</form>", "", cleaned)

    # 去掉 onload / onclick 等事件
    cleaned = re.sub(r'\son\w+="[^"]*"', "", cleaned, flags=re.I)
    cleaned = re.sub(r"\son\w+='[^']*'", "", cleaned, flags=re.I)

    # 屏蔽 javascript: 协议
    cleaned = re.sub(r'javascript:', "", cleaned, flags=re.I)

    return cleaned.strip()


def extract_body_content(msg, max_len=None):
    plain_parts = []
    html_parts_raw = []

    try:
        if msg.is_multipart():
            for part in msg.walk():
                ctype = part.get_content_type()
                disp = str(part.get("Content-Disposition", ""))

                if "attachment" in disp.lower():
                    continue

                if ctype not in ("text/plain", "text/html"):
                    continue

                payload = part.get_payload(decode=True)
                if not payload:
                    continue

                charset = part.get_content_charset()
                text = safe_decode_payload(payload, charset)

                if ctype == "text/plain":
                    plain_parts.append(text)
                elif ctype == "text/html":
                    html_parts_raw.append(text)
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                charset = msg.get_content_charset()
                text = safe_decode_payload(payload, charset)

                if msg.get_content_type() == "text/html":
                    html_parts_raw.append(text)
                else:
                    plain_parts.append(text)

    except Exception:
        pass

    plain_text = "\n".join(plain_parts).strip()
    html_raw = "\n".join(html_parts_raw).strip()
    html_text = html_to_text_keep_tables(html_raw) if html_raw else ""

    # 预览优先使用更完整的内容
    if len(html_text.strip()) > len(plain_text.strip()):
        merged_text = html_text
    else:
        merged_text = plain_text or html_text

    merged_text = re.sub(r"\s+\n", "\n", merged_text)
    merged_text = re.sub(r"\n{3,}", "\n\n", merged_text)

    if max_len:
        merged_text = merged_text[:max_len]

    return {
        "plain_text": plain_text,
        "html_raw": html_raw,
        "html_display": sanitize_html_for_display(html_raw),
        "merged_text": merged_text,
    }


def parse_email_date(date_str):
    try:
        dt = parsedate_to_datetime(date_str)
        if not dt:
            return None
        if dt.tzinfo:
            dt = dt.astimezone().replace(tzinfo=None)
        return dt
    except Exception:
        return None


def get_date_bucket(date_str):
    dt = parse_email_date(date_str)
    if not dt:
        return "unknown"

    now = datetime.now()
    if dt.date() == now.date():
        return "today"
    if dt >= now - timedelta(days=7):
        return "last7"
    return "older"


def format_datetime_for_sort(date_str):
    dt = parse_email_date(date_str)
    if dt:
        return dt
    return datetime.min


def classify_mail(text: str) -> str:
    t = (text or "").lower()
    for category, keywords in CATEGORY_RULES.items():
        for kw in keywords:
            if kw.lower() in t:
                return category
    return "其他风险类"


def get_hit_keywords(text: str, limit=8):
    t = text or ""
    hits = []
    for raw, rgx in zip(IMPORTANT_LIKE_RULES, RULE_REGEX):
        if rgx.search(t):
            clean_kw = raw.replace("%", "")
            if clean_kw not in hits:
                hits.append(clean_kw)
        if len(hits) >= limit:
            break
    return " | ".join(hits)


def is_important(text: str) -> bool:
    t = (text or "").lower()

    if any(r.search(t) for r in RULE_REGEX):
        return True

    broad_keywords = [
        "action required",
        "important",
        "urgent",
        "verification",
        "document required",
        "compliance",
        "product safety",
        "safety documentation",
        "listing deactivation",
        "listing deactivated",
        "reactivate listings",
        "listing removed",
        "removed",
        "seller account",
        "amazon account",
        "performance notification",
        "intellectual property",
        "infringement",
        "rights owner",
        "funds transfer",
        "disbursement",
        "payment",
        "auszahlung",
        "pago",
        "płatność",
        "hazard",
        "unsafe",
        "recall",
        "appeal",
        "kyc",
        "court",
        "tro",
        "plaintiff",
        "angebote wurden aus",
        "entfernt",
        "genehmigt wird",
        "removal order",
    ]
    return any(kw in t for kw in broad_keywords)


def save_xlsx_if_possible(rows, filename):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Important Mails"

        headers = [
            "category", "time_bucket", "account", "date", "from",
            "subject", "hit_keywords", "body_preview"
        ]
        ws.append(headers)

        for row in rows:
            ws.append([
                row.get("category", ""),
                row.get("time_bucket", ""),
                row.get("account", ""),
                row.get("date", ""),
                row.get("from", ""),
                row.get("subject", ""),
                row.get("hit_keywords", ""),
                row.get("body_preview", ""),
            ])

        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        bold_font = Font(bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        widths = {
            "A": 18,
            "B": 12,
            "C": 20,
            "D": 30,
            "E": 36,
            "F": 70,
            "G": 50,
            "H": 100,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        wb.save(filename)
        return True
    except Exception:
        return False


def fetch_message_safely(server, msg_num: int):
    try:
        resp, lines, octets = server.retr(msg_num)
        return b"\r\n".join(lines), "RETR"
    except Exception as e1:
        try:
            resp, lines, octets = server.top(msg_num, 120)
            return b"\r\n".join(lines), "TOP"
        except Exception as e2:
            raise RuntimeError(f"RETR失败: {e1}; TOP失败: {e2}")


def build_mail_anchor(row, idx):
    raw = f'{row.get("category","")}_{row.get("account","")}_{row.get("uid","")}_{idx}'
    raw = re.sub(r"[^a-zA-Z0-9_-]+", "_", raw)
    return f"mail_{raw}"


def group_rows_by_category(rows):
    grouped = {}
    for row in rows:
        category = row.get("category", "其他风险类")
        grouped.setdefault(category, []).append(row)
    return grouped


def filter_rows_by_time(rows, mode):
    if mode == "today":
        return [r for r in rows if r.get("time_bucket") == "today"]
    if mode == "last7":
        return [r for r in rows if r.get("time_bucket") in ("today", "last7")]
    return rows[:]


def count_rows_by_bucket(rows):
    today_count = sum(1 for r in rows if r.get("time_bucket") == "today")
    last7_count = sum(1 for r in rows if r.get("time_bucket") in ("today", "last7"))
    all_count = len(rows)
    return today_count, last7_count, all_count


def dedup_and_prepare_rows(rows):
    dedup = {}
    for row in rows:
        key = (
            row.get("account", ""),
            row.get("subject", ""),
            row.get("date", ""),
            row.get("uid", "")
        )
        clean = dict(row)
        if "time_bucket" not in clean:
            clean["time_bucket"] = get_date_bucket(clean.get("date", ""))
        clean["_sort_dt"] = format_datetime_for_sort(clean.get("date", ""))
        dedup[key] = clean

    history_rows = list(dedup.values())

    category_rank = {
        "店铺绩效类": 1,
        "KYC审核类": 2,
        "付款/资金类": 3,
        "产品链接/合规类": 4,
        "法务/侵权类": 5,
        "货件/库存类": 6,
        "其他风险类": 7,
    }
    bucket_rank = {
        "today": 1,
        "last7": 2,
        "older": 3,
        "unknown": 4,
    }

    history_rows.sort(
        key=lambda x: (
            bucket_rank.get(x.get("time_bucket", "unknown"), 99),
            category_rank.get(x.get("category", "其他风险类"), 99),
            -x.get("_sort_dt", datetime.min).timestamp()
            if x.get("_sort_dt") and x.get("_sort_dt") != datetime.min else 0
        )
    )

    for row in history_rows:
        row.pop("_sort_dt", None)

    return history_rows


def generate_html_report(rows, account_summary, report_time_str, html_filename, title):
    hit_today_count, hit_last7_count, hit_all_count = count_rows_by_bucket(rows)

    total_today_mail = sum(x.get("today_mail_count", 0) for x in account_summary)
    total_last7_mail = sum(x.get("last7_mail_count", 0) for x in account_summary)
    total_all_mail = sum(x.get("all_mail_count", 0) for x in account_summary)
    total_skipped = sum(x.get("skipped_count", 0) for x in account_summary)
    login_failed_count = sum(1 for x in account_summary if str(x.get("login_status", "")).startswith("失败"))

    category_order = [
        "店铺绩效类",
        "KYC审核类",
        "付款/资金类",
        "产品链接/合规类",
        "法务/侵权类",
        "货件/库存类",
        "其他风险类",
    ]

    def build_category_summary(mode):
        sub_rows = filter_rows_by_time(rows, mode)
        grouped = group_rows_by_category(sub_rows)
        html_rows = ""
        for cat in category_order:
            count = len(grouped.get(cat, []))
            html_rows += f"""
            <tr>
                <td>{html.escape(cat)}</td>
                <td>{count}</td>
            </tr>
            """
        return html_rows

    category_summary_today = build_category_summary("today")
    category_summary_last7 = build_category_summary("last7")
    category_summary_all = build_category_summary("all")

    account_summary_html = ""
    for item in account_summary:
        account_summary_html += f"""
        <tr>
            <td>{html.escape(item.get("account", ""))}</td>
            <td>{html.escape(item.get("user", ""))}</td>
            <td>{item.get("today_mail_count", 0)}</td>
            <td>{item.get("last7_mail_count", 0)}</td>
            <td>{item.get("all_mail_count", 0)}</td>
            <td>{item.get("hit_count", 0)}</td>
            <td>{item.get("skipped_count", 0)}</td>
            <td>{html.escape(item.get("login_status", ""))}</td>
        </tr>
        """

    def build_sections(mode):
        sub_rows = filter_rows_by_time(rows, mode)
        grouped = group_rows_by_category(sub_rows)

        category_sections_html = ""
        for cat in category_order:
            cat_rows = grouped.get(cat, [])
            if not cat_rows:
                continue

            cards_html = ""
            for idx, row in enumerate(cat_rows, 1):
                anchor = build_mail_anchor(row, idx)
                subject = html.escape(row.get("subject", ""))
                from_ = html.escape(row.get("from", ""))
                date_ = html.escape(row.get("date", ""))
                account = html.escape(row.get("account", ""))
                hits = html.escape(row.get("hit_keywords", ""))
                preview = html.escape(row.get("body_preview", ""))
                fetch_mode = html.escape(row.get("fetch_mode", ""))
                detected_at = html.escape(row.get("detected_at", ""))
                time_bucket = html.escape(row.get("time_bucket", ""))
                html_body = row.get("html_body", "") or ""

                cards_html += f"""
                <div class="mail-card searchable-card" data-search="{(subject + ' ' + from_ + ' ' + hits + ' ' + account).lower()}">
                    <div class="mail-title" id="{anchor}">{subject}</div>
                    <div class="mail-meta">
                        <span><b>时间分组：</b>{time_bucket}</span>
                        <span><b>邮箱：</b>{account}</span>
                        <span><b>发件人：</b>{from_}</span>
                        <span><b>日期：</b>{date_}</span>
                        <span><b>命中词：</b>{hits}</span>
                        <span><b>抓取方式：</b>{fetch_mode}</span>
                        <span><b>检测时间：</b>{detected_at}</span>
                    </div>

                    <details>
                        <summary>展开查看邮件预览</summary>
                        <pre>{preview}</pre>
                    </details>
                """

                if html_body.strip():
                    cards_html += f"""
                    <details class="html-details">
                        <summary>展开查看原始表格内容</summary>
                        <div class="raw-html-box">
                            {html_body}
                        </div>
                    </details>
                    """

                cards_html += "</div>"

            category_sections_html += f"""
            <div class="category-block time-panel" data-mode="{mode}">
                <div class="category-header">
                    <div class="category-name">{html.escape(cat)}</div>
                    <div class="category-count">共 {len(cat_rows)} 封</div>
                </div>
                <details class="category-details">
                    <summary>展开/收起</summary>
                    <div class="category-body">
                        {cards_html}
                    </div>
                </details>
            </div>
            """

        if not category_sections_html:
            category_sections_html = f"""
            <div class="empty-box time-panel" data-mode="{mode}">当前筛选下没有命中风险邮件</div>
            """

        return category_sections_html

    sections_today = build_sections("today")
    sections_last7 = build_sections("last7")
    sections_all = build_sections("all")

    html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>{html.escape(title)}</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        * {{
            box-sizing: border-box;
        }}
        body {{
            margin: 0;
            background: #f5f7fb;
            color: #222;
            font-family: Arial, "Microsoft YaHei", sans-serif;
        }}
        .container {{
            width: 96%;
            max-width: 1500px;
            margin: 20px auto;
        }}
        .top-bar {{
            background: linear-gradient(135deg, #1f3a8a, #2563eb);
            color: #fff;
            padding: 18px 22px;
            border-radius: 14px;
            margin-bottom: 16px;
            box-shadow: 0 8px 24px rgba(37, 99, 235, 0.18);
        }}
        .top-title {{
            font-size: 26px;
            font-weight: bold;
            margin-bottom: 8px;
        }}
        .top-desc {{
            font-size: 14px;
            line-height: 1.8;
            opacity: .95;
        }}
        .toolbar {{
            display: flex;
            flex-wrap: wrap;
            gap: 10px;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 14px;
        }}
        .tab-group {{
            display: flex;
            flex-wrap: wrap;
            gap: 10px;
        }}
        .tab-btn {{
            border: 1px solid #cbd5e1;
            background: #fff;
            padding: 10px 16px;
            border-radius: 10px;
            cursor: pointer;
            font-size: 14px;
            font-weight: bold;
        }}
        .tab-btn.active {{
            background: #2563eb;
            color: #fff;
            border-color: #2563eb;
        }}
        .search-box {{
            min-width: 260px;
            flex: 1;
            max-width: 420px;
        }}
        .search-box input {{
            width: 100%;
            padding: 11px 14px;
            border: 1px solid #cbd5e1;
            border-radius: 10px;
            font-size: 14px;
        }}
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(7, 1fr);
            gap: 12px;
            margin-bottom: 16px;
        }}
        .summary-box {{
            background: #fff;
            border: 1px solid #dbe2ea;
            border-radius: 12px;
            padding: 16px;
            box-shadow: 0 4px 10px rgba(15, 23, 42, 0.04);
        }}
        .summary-label {{
            color: #667085;
            font-size: 13px;
            margin-bottom: 8px;
        }}
        .summary-value {{
            font-size: 28px;
            font-weight: bold;
            color: #0f172a;
        }}
        .panel {{
            background: #fff;
            border: 1px solid #dbe2ea;
            border-radius: 12px;
            margin-bottom: 16px;
            overflow: hidden;
            box-shadow: 0 4px 10px rgba(15, 23, 42, 0.04);
        }}
        .panel-title {{
            font-size: 16px;
            font-weight: bold;
            padding: 12px 16px;
            border-bottom: 1px solid #ebeef5;
            background: #f8fafc;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th, td {{
            border: 1px solid #ebeef5;
            padding: 10px 12px;
            text-align: left;
            font-size: 14px;
            vertical-align: top;
            word-break: break-word;
        }}
        th {{
            background: #f8fafc;
            font-weight: bold;
        }}
        .category-block {{
            background: #fff;
            border: 1px solid #dbe2ea;
            border-radius: 12px;
            margin-bottom: 16px;
            overflow: hidden;
            box-shadow: 0 4px 10px rgba(15, 23, 42, 0.04);
        }}
        .category-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            background: #f8fafc;
            border-bottom: 1px solid #ebeef5;
            padding: 12px 16px;
        }}
        .category-name {{
            font-size: 18px;
            font-weight: bold;
        }}
        .category-count {{
            color: #64748b;
            font-size: 14px;
        }}
        .category-details > summary {{
            cursor: pointer;
            padding: 12px 16px;
            background: #fff;
            font-weight: bold;
            border-bottom: 1px solid #f1f5f9;
        }}
        .category-body {{
            padding: 14px;
        }}
        .mail-card {{
            border: 1px solid #e2e8f0;
            background: #fcfcfd;
            border-radius: 10px;
            padding: 14px;
            margin-bottom: 12px;
        }}
        .mail-title {{
            font-size: 16px;
            font-weight: bold;
            color: #111827;
            margin-bottom: 10px;
            line-height: 1.6;
        }}
        .mail-meta {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 8px 18px;
            font-size: 13px;
            color: #334155;
            margin-bottom: 10px;
        }}
        details {{
            background: #fff;
            border: 1px solid #e2e8f0;
            border-radius: 8px;
            padding: 10px 12px;
            margin-top: 10px;
        }}
        summary {{
            cursor: pointer;
            font-weight: bold;
        }}
        pre {{
            white-space: pre-wrap;
            word-break: break-word;
            margin: 10px 0 0 0;
            font-family: Arial, "Microsoft YaHei", sans-serif;
            font-size: 13px;
            line-height: 1.65;
        }}
        .raw-html-box {{
            background: #fff;
            border: 1px solid #ddd;
            padding: 12px;
            margin-top: 10px;
            overflow-x: auto;
            max-width: 100%;
        }}
        .raw-html-box table {{
            border-collapse: collapse;
            width: auto;
            min-width: 600px;
            max-width: none;
        }}
        .raw-html-box th,
        .raw-html-box td {{
            border: 1px solid #999;
            padding: 8px;
            text-align: left;
            vertical-align: top;
            font-size: 13px;
            line-height: 1.5;
            word-break: break-word;
        }}
        .raw-html-box img {{
            max-width: 100%;
            height: auto;
        }}
        .empty-box {{
            background: #fff;
            border: 1px solid #dbe2ea;
            border-radius: 12px;
            padding: 30px;
            text-align: center;
            color: #64748b;
        }}
        .hidden {{
            display: none !important;
        }}
        @media (max-width: 1200px) {{
            .summary-grid {{
                grid-template-columns: repeat(4, 1fr);
            }}
        }}
        @media (max-width: 900px) {{
            .mail-meta {{
                grid-template-columns: 1fr;
            }}
            .summary-grid {{
                grid-template-columns: repeat(2, 1fr);
            }}
        }}
        @media (max-width: 600px) {{
            .summary-grid {{
                grid-template-columns: 1fr;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="top-bar">
            <div class="top-title">{html.escape(title)}</div>
            <div class="top-desc">
                生成时间：{html.escape(report_time_str)}<br>
                当前版本按邮件日期统计：今日 / 最近7天 / 全部，不再依赖“新邮件”标记。
            </div>
        </div>

        <div class="toolbar">
            <div class="tab-group">
                <button class="tab-btn active" data-mode="today">今日命中（{hit_today_count}）</button>
                <button class="tab-btn" data-mode="last7">最近7天命中（{hit_last7_count}）</button>
                <button class="tab-btn" data-mode="all">全部命中（{hit_all_count}）</button>
            </div>
            <div class="search-box">
                <input id="searchInput" type="text" placeholder="搜索主题 / 发件人 / 命中词 / 邮箱">
            </div>
        </div>

        <div class="summary-grid">
            <div class="summary-box">
                <div class="summary-label">今日邮件总数</div>
                <div class="summary-value">{total_today_mail}</div>
            </div>
            <div class="summary-box">
                <div class="summary-label">最近7天邮件总数</div>
                <div class="summary-value">{total_last7_mail}</div>
            </div>
            <div class="summary-box">
                <div class="summary-label">全部邮件总数</div>
                <div class="summary-value">{total_all_mail}</div>
            </div>
            <div class="summary-box">
                <div class="summary-label">今日命中</div>
                <div class="summary-value">{hit_today_count}</div>
            </div>
            <div class="summary-box">
                <div class="summary-label">最近7天命中</div>
                <div class="summary-value">{hit_last7_count}</div>
            </div>
            <div class="summary-box">
                <div class="summary-label">全部命中</div>
                <div class="summary-value">{hit_all_count}</div>
            </div>
            <div class="summary-box">
                <div class="summary-label">登录失败邮箱数</div>
                <div class="summary-value">{login_failed_count}</div>
            </div>
        </div>

        <div class="panel">
            <div class="panel-title">分类汇总</div>

            <div class="time-panel" data-mode="today">
                <table>
                    <thead>
                        <tr><th>分类</th><th>命中数量</th></tr>
                    </thead>
                    <tbody>
                        {category_summary_today}
                    </tbody>
                </table>
            </div>

            <div class="time-panel hidden" data-mode="last7">
                <table>
                    <thead>
                        <tr><th>分类</th><th>命中数量</th></tr>
                    </thead>
                    <tbody>
                        {category_summary_last7}
                    </tbody>
                </table>
            </div>

            <div class="time-panel hidden" data-mode="all">
                <table>
                    <thead>
                        <tr><th>分类</th><th>命中数量</th></tr>
                    </thead>
                    <tbody>
                        {category_summary_all}
                    </tbody>
                </table>
            </div>
        </div>

        <div class="panel">
            <div class="panel-title">邮箱汇总（本次运行）</div>
            <table>
                <thead>
                    <tr>
                        <th>账户名称</th>
                        <th>邮箱</th>
                        <th>今日邮件</th>
                        <th>最近7天邮件</th>
                        <th>全部邮件</th>
                        <th>命中</th>
                        <th>跳过</th>
                        <th>登录状态</th>
                    </tr>
                </thead>
                <tbody>
                    {account_summary_html}
                </tbody>
            </table>
        </div>

        {sections_today}
        {sections_last7}
        {sections_all}
    </div>

    <script>
        const tabButtons = document.querySelectorAll('.tab-btn');
        const timePanels = document.querySelectorAll('.time-panel');
        const searchInput = document.getElementById('searchInput');

        function switchMode(mode) {{
            tabButtons.forEach(btn => {{
                btn.classList.toggle('active', btn.dataset.mode === mode);
            }});
            timePanels.forEach(panel => {{
                panel.classList.toggle('hidden', panel.dataset.mode !== mode);
            }});
            applySearch();
        }}

        function applySearch() {{
            const kw = (searchInput.value || '').trim().toLowerCase();
            const visiblePanels = document.querySelectorAll('.time-panel:not(.hidden) .searchable-card');

            document.querySelectorAll('.searchable-card').forEach(card => {{
                card.classList.remove('hidden');
            }});

            if (!kw) return;

            visiblePanels.forEach(card => {{
                const hay = card.dataset.search || '';
                if (!hay.includes(kw)) {{
                    card.classList.add('hidden');
                }}
            }});
        }}

        tabButtons.forEach(btn => {{
            btn.addEventListener('click', () => switchMode(btn.dataset.mode));
        }});

        searchInput.addEventListener('input', applySearch);

        switchMode('today');
    </script>
</body>
</html>
"""
    with open(html_filename, "w", encoding="utf-8") as f:
        f.write(html_content)


def login_with_retry(host, user, password, retries=3, delay=4):
    last_error = None
    for attempt in range(1, retries + 1):
        server = None
        try:
            print(f"尝试登录，第 {attempt}/{retries} 次：{user} @ {host}")
            server = poplib.POP3_SSL(host, 995, timeout=30)
            server.user(user)
            server.pass_(password)
            return server
        except Exception as e:
            last_error = e
            print(f"登录失败，第 {attempt} 次：{e}")
            try:
                if server:
                    server.quit()
            except Exception:
                pass
            if attempt < retries:
                print(f"等待 {delay} 秒后重试...")
                time.sleep(delay)

    raise Exception(f"多次尝试仍无法登录：{last_error}")


# =========================
# 4) 主程序
# =========================
def main():
    accounts = load_accounts_from_env()

    # 打乱顺序，降低固定顺序触发风控的概率
    random.shuffle(accounts)

    print("\n===== 实际读取到的邮箱配置 =====")
    for acct in accounts:
        print(
            f'{acct["name"]} | {acct["user"]} | {acct["host"]} | 密码长度={len(acct["password"])}'
        )
    print("=" * 60)

    today_str = datetime.now().strftime("%Y-%m-%d")
    report_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    important_rows = []
    account_summary = []

    for acct in accounts:
        name = acct.get("name", acct.get("user", ""))
        host = acct["host"]
        user = acct["user"]
        password = acct["password"]

        print("\n" + "=" * 60)
        print(f"检查邮箱：{name} ({user})")
        print("=" * 60)

        server = None

        try:
            print(f"准备连接 HOST: {host}")
            print(f"登录邮箱: {user}")
            print(f"密码长度: {len(password)}")

            server = login_with_retry(host, user, password, retries=3, delay=4)
            print(f"[{name}] 登录成功")

            resp, uid_lines, octets = server.uidl()
            uid_map = {}
            for line in uid_lines:
                parts = line.decode("utf-8", errors="replace").split()
                if len(parts) >= 2:
                    msg_num = parts[0]
                    uid = parts[1]
                    uid_map[msg_num] = uid

            # 按邮件序号倒序处理，新的在前
            all_msg_nums = sorted(uid_map.keys(), key=lambda x: int(x), reverse=True)

            print(f"邮箱当前可读取邮件 {len(all_msg_nums)} 封")

            found_count = 0
            skipped_count = 0

            today_mail_count = 0
            last7_mail_count = 0
            all_mail_count = 0

            for msg_num in all_msg_nums:
                uid = uid_map[msg_num]

                try:
                    raw, fetch_mode = fetch_message_safely(server, int(msg_num))
                    msg = email.message_from_bytes(raw)

                    subject, from_, date_ = get_text_headers(msg)
                    date_bucket = get_date_bucket(date_)

                    all_mail_count += 1
                    if date_bucket == "today":
                        today_mail_count += 1
                    if date_bucket in ("today", "last7"):
                        last7_mail_count += 1

                    body_data = extract_body_content(msg, max_len=20000)
                    body_full = body_data.get("merged_text", "")
                    body_preview = (body_full or "")[:1200]
                    html_body = body_data.get("html_display", "")

                    haystack = f"{subject}\n{from_}\n{date_}\n{body_full}"

                    if is_important(haystack):
                        row = {
                            "category": classify_mail(haystack),
                            "time_bucket": date_bucket,
                            "account": name,
                            "account_user": user,
                            "date": date_,
                            "from": from_,
                            "subject": subject,
                            "hit_keywords": get_hit_keywords(haystack),
                            "body_preview": body_preview,
                            "body_full": body_full,
                            "html_body": html_body,
                            "uid": uid,
                            "detected_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "fetch_mode": fetch_mode,
                            "_sort_dt": format_datetime_for_sort(date_)
                        }
                        important_rows.append(row)
                        found_count += 1

                except Exception as e:
                    skipped_count += 1
                    print(f"跳过邮件 {msg_num}，原因：{e}")

            print(
                f"[{name}] 今日邮件 {today_mail_count} 封，"
                f"最近7天邮件 {last7_mail_count} 封，"
                f"全部邮件 {all_mail_count} 封，"
                f"命中 {found_count} 封，跳过 {skipped_count} 封"
            )

            account_summary.append({
                "account": name,
                "user": user,
                "today_mail_count": today_mail_count,
                "last7_mail_count": last7_mail_count,
                "all_mail_count": all_mail_count,
                "hit_count": found_count,
                "skipped_count": skipped_count,
                "login_status": "成功",
            })

        except Exception as e:
            print(f"[{name}] 邮箱连接失败：{e}")
            account_summary.append({
                "account": name,
                "user": user,
                "today_mail_count": 0,
                "last7_mail_count": 0,
                "all_mail_count": 0,
                "hit_count": 0,
                "skipped_count": 0,
                "login_status": f"失败：{e}",
            })

        finally:
            if server:
                try:
                    server.quit()
                except Exception:
                    pass

            print("等待 3 秒，准备检查下一个邮箱...")
            time.sleep(3)

    final_rows = dedup_and_prepare_rows(important_rows)

    print("\n" + "=" * 60)
    print("所有邮箱检测汇总")
    print("=" * 60)

    total_today_mail = 0
    total_last7_mail = 0
    total_all_mail = 0
    total_hit = 0
    total_skipped = 0

    for item in account_summary:
        total_today_mail += item["today_mail_count"]
        total_last7_mail += item["last7_mail_count"]
        total_all_mail += item["all_mail_count"]
        total_hit += item["hit_count"]
        total_skipped += item["skipped_count"]
        print(
            f'{item["account"]} ({item["user"]})：'
            f'今日邮件 {item["today_mail_count"]} 封，'
            f'最近7天邮件 {item["last7_mail_count"]} 封，'
            f'全部邮件 {item["all_mail_count"]} 封，'
            f'命中 {item["hit_count"]} 封，'
            f'跳过 {item["skipped_count"]} 封，'
            f'登录状态：{item.get("login_status", "")}'
        )

    print("-" * 60)
    print(f"今日邮件总数：{total_today_mail} 封")
    print(f"最近7天邮件总数：{total_last7_mail} 封")
    print(f"全部邮件总数：{total_all_mail} 封")
    print(f"全部命中：{total_hit} 封")
    print(f"总跳过：{total_skipped} 封")

    html_name = f"important_mails_{today_str}.html"
    csv_name = f"important_mails_{today_str}.csv"
    xlsx_name = f"important_mails_{today_str}.xlsx"
    json_name = f"important_mails_{today_str}.json"

    with open(csv_name, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "category", "time_bucket", "account", "account_user", "date", "from",
                "subject", "hit_keywords", "body_preview",
                "detected_at", "fetch_mode", "uid"
            ]
        )
        writer.writeheader()
        for row in final_rows:
            writer.writerow({
                "category": row.get("category", ""),
                "time_bucket": row.get("time_bucket", ""),
                "account": row.get("account", ""),
                "account_user": row.get("account_user", ""),
                "date": row.get("date", ""),
                "from": row.get("from", ""),
                "subject": row.get("subject", ""),
                "hit_keywords": row.get("hit_keywords", ""),
                "body_preview": row.get("body_preview", ""),
                "detected_at": row.get("detected_at", ""),
                "fetch_mode": row.get("fetch_mode", ""),
                "uid": row.get("uid", ""),
            })

    with open(json_name, "w", encoding="utf-8") as f:
        json.dump(final_rows, f, ensure_ascii=False, indent=2)

    xlsx_ok = save_xlsx_if_possible(final_rows, xlsx_name)

    generate_html_report(
        rows=final_rows,
        account_summary=account_summary,
        report_time_str=report_time_str,
        html_filename=html_name,
        title=f"风险邮件后台报告 - {today_str}",
    )

    print("-" * 60)
    if final_rows:
        print(f"本次扫描命中风险邮件 {len(final_rows)} 封")
    else:
        print("本次没有命中风险邮件")
    print(f"已生成：{csv_name}")
    print(f"已生成：{json_name}")
    print(f"已生成：{html_name}")
    if xlsx_ok:
        print(f"已生成：{xlsx_name}")
    else:
        print("未生成 xlsx（先运行：pip install openpyxl）")

    print("=" * 60)

    src = html_name
    dst = "index.html"

    if os.path.exists(src):
        shutil.copy(src, dst)
        print(f"已同步网页首页: {src} -> {dst}")
    else:
        print(f"未找到HTML文件，无法复制: {src}")


if __name__ == "__main__":
    main()